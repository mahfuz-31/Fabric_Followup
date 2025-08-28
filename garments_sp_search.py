import pandas as pd
import requests as rq
from bs4 import BeautifulSoup as bs

def garments_sp_search(root, progress, orders):
    progress.pack(pady=10)
    progress.pack_configure(anchor='center')
    progress.stop()  # in case it was spinning before
    orders = list(set(orders))
    orders = [o.strip() for o in orders if o.strip()]
    no_of_orders = len(orders)
    if no_of_orders == 0:
        progress.pack_forget()
        return
    progress.config(mode='determinate', maximum=no_of_orders, value=0)
    root.update_idletasks()

    result_df = pd.DataFrame()
    columns = ['Order No',	'Buyer',	'Order Qty (Pcs)',	'Cutting Qty (Pcs)',	'Cutting Rejection (Pcs)',	'Panel Rejection',	'Input Qty (Pcs)',	'Sewing Rejection (Pcs)',	'Output Qty (Pcs)',	'Finishing Rejection (Pcs)',	'Poly Qty (Pcs)',	'Leftover/Rej Qty (Pcs)',	'Floor Shipped Qty (Pcs)',	'Ex-Factory Shipped Qty (Pcs)', 'Prod Floor']
    result_df = result_df.reindex(columns=result_df.columns.tolist() + columns)

    idx = 0
    for order in orders:
        url = 'http://192.168.1.150/MIS365/PHPReports/dbl/view.php?page=pro_details&m=Order%20Wise%20Garments%20Status&w=900&OrderNo=' + str(order)
        
        response = rq.get(url)
        html_content = bs(response.content, 'html.parser')
        rows = []
        for row in html_content.find_all('tr'):
            row_data = [cell.get_text(strip=True) for cell in row.find_all('td')]
            rows.append(row_data)
        result_row = rows[len(rows) - 1]
        i = 0
        for col in columns:
            if i > len(result_row) - 1:
                break
            elif i > 1 and result_row[i] != '':
                result_df.loc[idx, col] = int(result_row[i])
            else:
                result_df.loc[idx, col] = result_row[i]
            i += 1
        
        
        inp_url = 'http://192.168.1.150/MIS365/PHPReports/dbl/view.php?page=outputNew&m=Unit%20wise%20Output%20Information&w=900&OrderNo=' + str(order)
        response = rq.get(inp_url)
        html_content = bs(response.content, 'html.parser')
        rows = []
        for row in html_content.find_all('tr'):
            row_data = [cell.get_text(strip=True) for cell in row.find_all('td')]
            rows.append(row_data)
        for row in rows:
            if len(row) == 5:
                result_df.loc[idx, 'Prod Floor'] = row[1]
        idx += 1
        progress['value'] += idx
        root.update_idletasks()
    try:
        result_df.to_excel('garments_status_output.xlsx', index=False)

        from openpyxl import load_workbook
        from openpyxl.styles import Border, Side, Alignment

        wb = load_workbook('garments_status_output.xlsx')

        ws = wb['Sheet1']

        border = Border(left=Side(style='thin', color='0c8748'),
                        right=Side(style='thin', color='0c8748'),
                        top=Side(style='thin', color='0c8748'),
                        bottom=Side(style='thin', color='0c8748'))

        align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = align
        wb.save('garments_status_output.xlsx')
    except Exception as e:
        import tkinter.messagebox as msgbox
        msgbox.showerror("Error saving excel file: ", str(e))
    progress['value'] = no_of_orders
    root.update_idletasks()
    progress.stop()