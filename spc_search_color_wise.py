# Modified spc_search_color_wise.py for better GUI responsiveness
import requests as rq
from bs4 import BeautifulSoup as bs
import pandas as pd
import time

def convert_to_number(s):
    # Remove commas from the string
    s = s.replace(",", "")
    # Convert to float if there's a decimal point, otherwise to convert_to_number
    return float(s) if '.' in s else int(s)
def spc_search_color_wise(root, progress, orders):
    # Show progress bar in DETERMINATE mode
    progress.pack(pady=10)
    progress.pack_configure(anchor='center')
    progress.stop()  # in case it was spinning before
    orders = list(set(orders))
    orders = [o.strip() for o in orders if o.strip()]
    total = len(orders)

    if total == 0:
        progress.pack_forget()
        return

    # Configure determinate progress
    progress.config(mode='determinate', maximum=total, value=0)
    root.update_idletasks()

    print("Orders:", orders)

    df = pd.DataFrame()
    columns = ["Buyer", "Order", 'Style', 'UoF', 'F. Color', 'G. Color', 'Y. Type', 'F. Type', 'GSM', 'Dia',
               'G/F Order With S.Note Qty', 'G/F S.Note Qty', 'Net Grey Receive Qty', 'G/F Rcv Balance Qty',
               'F/F Order with S.Note Qty', 'F/F S.Note Qty', 'F/F Delv Qty', 'F/F Delv. Balance Qty',
               'Replacement Delivery', 'F/F Excess Delv.Qty', 'Transfer To', 'Transfer From', 'Return Receive',
               'Return Delivery', 'Dyeing Unit', 'Order Sheet Receive Date', 'Cut Plan Start Date',
               'Cut Plan End Date', 'Delivery Unit']
    df = df.reindex(columns=df.columns.tolist() + columns)

    row_idx = 0
    for idx, order in enumerate(orders, start=1):
        try:
            url = f'http://192.168.13.253/mymun/Work%20Order/combineSearchResult.php?Welcome=7&GetOrderNO={order}'
            response = rq.get(url, timeout=30)

            html_content = bs(response.content, 'html.parser')
            rows = []
            for row in html_content.find_all('tr'):
                row_data = [cell.get_text(strip=True) for cell in row.find_all('td')]
                rows.append(row_data)

            row_idx2 = row_idx
            si = 1

            # --- Grey fabric ---
            for row in rows:
                if len(row) > 0 and row[0].isnumeric() and row[0] != '0':
                    if int(row[0]) < si:
                        break
                    if convert_to_number(row[8]) == 0:
                        continue

                    df.loc[row_idx, 'Buyer'] = rows[3][0] if len(rows) > 3 else ""
                    df.loc[row_idx, 'Order'] = rows[3][1] if len(rows) > 3 and len(rows[3]) > 1 else ""
                    df.loc[row_idx, 'Style'] = rows[3][2] if len(rows) > 3 and len(rows[3]) > 2 else ""
                    df.loc[row_idx, 'UoF'] = row[1] if len(row) > 1 else ""
                    df.loc[row_idx, 'F. Color'] = row[2][0:row[2].find('(')] if len(row) > 2 and '(' in row[2] else (row[2] if len(row) > 2 else "")
                    df.loc[row_idx, 'G. Color'] = row[3] if len(row) > 3 else ""
                    df.loc[row_idx, 'Y. Type'] = row[4] if len(row) > 4 else ""
                    df.loc[row_idx, 'F. Type'] = row[5] if len(row) > 5 else ""
                    df.loc[row_idx, 'GSM'] = convert_to_number(row[6]) if len(row) > 6 and row[6] != "" else ""
                    df.loc[row_idx, 'Dia'] = row[7] if len(row) > 7 else ""
                    df.loc[row_idx, 'G/F Order With S.Note Qty'] = convert_to_number(row[8]) if len(row) > 8 else 0
                    df.loc[row_idx, 'G/F S.Note Qty'] = convert_to_number(row[9]) if len(row) > 9 else 0
                    df.loc[row_idx, 'Net Grey Receive Qty'] = convert_to_number(row[13]) if len(row) > 13 else 0
                    df.loc[row_idx, 'G/F Rcv Balance Qty'] = convert_to_number(row[15]) if len(row) > 15 else 0
                    row_idx += 1
                    si += 1

            # --- Finish fabric ---
            row_idx = row_idx2
            si = 0
            for row in rows:
                if len(row) == 1 and row[0].endswith('Finish Fabric Delivery'):
                    if row[0] == 'Finish Fabric Delivery':
                        df.loc[row_idx, 'Delivery Unit'] = 'Jinnat Complex'
                    else:
                        unit = row[0].split('::')[1].split('||')[0]
                        df.loc[row_idx, 'Delivery Unit'] = unit.strip()

                if len(row) > 0 and row[0] == 'SL NO.':
                    si += 1

                if len(row) > 0 and si == 2 and row[0].isnumeric() and row[0] != '0':
                    if len(row) > 8 and convert_to_number(row[8]) == 0:
                        continue

                    if len(row) > 8:
                        df.loc[row_idx, 'F/F Order with S.Note Qty'] = convert_to_number(row[8])
                    if len(row) > 9:
                        df.loc[row_idx, 'F/F S.Note Qty'] = convert_to_number(row[9])
                    if len(row) > 12:
                        df.loc[row_idx, 'F/F Delv Qty'] = convert_to_number(row[12])
                    if len(row) > 15:
                        df.loc[row_idx, 'F/F Delv. Balance Qty'] = convert_to_number(row[15])
                    if len(row) > 16:
                        df.loc[row_idx, 'Replacement Delivery'] = convert_to_number(row[16])
                    if len(row) > 18:
                        df.loc[row_idx, 'F/F Excess Delv.Qty'] = convert_to_number(row[18])
                    if len(row) > 19:
                        df.loc[row_idx, 'Transfer To'] = row[19]
                    if len(row) > 20:
                        df.loc[row_idx, 'Transfer From'] = row[20]
                    if len(row) > 22:
                        df.loc[row_idx, 'Return Receive'] = convert_to_number(row[22])
                    if len(row) > 23:
                        df.loc[row_idx, 'Return Delivery'] = convert_to_number(row[23])
                    if len(row) > 26:
                        df.loc[row_idx, 'Dyeing Unit'] = row[26]

                    if len(rows) > 3:
                        df.loc[row_idx, 'Order Sheet Receive Date'] = rows[3][4] if len(rows[3]) > 4 else ""
                        df.loc[row_idx, 'Cut Plan Start Date'] = rows[3][5] if len(rows[3]) > 5 else ""
                        df.loc[row_idx, 'Cut Plan End Date'] = rows[3][6] if len(rows[3]) > 6 else ""

                    row_idx += 1

        except Exception as e:
            print(f"Error processing order {order}: {str(e)}")

        # <-- Progress increment after finishing this order
        progress['value'] = idx
        root.update_idletasks()

    # Save to Excel (unchanged)
    try:
        df.to_excel('color_wise_output.xlsx', index=False)
        from openpyxl import load_workbook
        from openpyxl.styles import Border, Side, Alignment

        wb = load_workbook('color_wise_output.xlsx')
        ws = wb['Sheet1']

        border = Border(left=Side(style='thin', color='0c8748'),
                        right=Side(style='thin', color='0c8748'),
                        top=Side(style='thin', color='0c8748'),
                        bottom=Side(style='thin', color='0c8748'))
        align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = align

        wb.save('color_wise_output.xlsx')
        print("Excel file saved successfully!")
    except Exception as e:
        print(f"Error saving Excel file: {str(e)}")

    # Done
    progress['value'] = total
    root.update_idletasks()
    progress.stop()
    # progress.pack_forget()
    